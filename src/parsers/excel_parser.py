from __future__ import annotations

from openpyxl import load_workbook
from loguru import logger

from src.models import TestCase


def parse_excel(path: str, url: str) -> list[TestCase]:
    """Parse an Excel file into TestCase list.

    Expected format: first row is headers (field names),
    each subsequent row is a test case.
    Columns named 'test_id', 'url', 'description', 'expected_outcome'
    are treated as metadata; all others are test data fields.
    """
    wb = load_workbook(path, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        raise ValueError(f"Excel file '{path}' has no data rows")

    headers = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(rows[0])]
    meta_keys = {"test_id", "url", "description", "expected_outcome"}
    test_cases: list[TestCase] = []

    for row_idx, row in enumerate(rows[1:], start=2):
        meta: dict[str, str] = {}
        data: dict[str, str] = {}
        for col_idx, val in enumerate(row):
            if col_idx >= len(headers):
                break
            key = headers[col_idx]
            str_val = str(val).strip() if val is not None else ""
            if not str_val:
                continue
            if key in meta_keys:
                meta[key] = str_val
            else:
                data[key] = str_val

        if not data:
            logger.warning(f"Row {row_idx} has no data fields, skipping")
            continue

        test_cases.append(TestCase(
            test_id=meta.get("test_id", f"excel_row_{row_idx}"),
            url=meta.get("url", url),
            data=data,
            description=meta.get("description", ""),
            expected_outcome=meta.get("expected_outcome", "success"),
        ))

    wb.close()
    logger.info(f"Parsed {len(test_cases)} test cases from {path}")
    return test_cases
